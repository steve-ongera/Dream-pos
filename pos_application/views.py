from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from decimal import Decimal
import json
from datetime import datetime, timedelta
from django.db import models
from .models import Product, Category, Sale, SaleItem, Customer, Discount, Inventory
from django.views.decorators.http import require_http_methods
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.http import JsonResponse
import json
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


@csrf_protect
def login_view(request):
    """
    Handle user login with professional form
    """
    if request.user.is_authenticated:
        return redirect('dashboard')  # Redirect to your main dashboard
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        remember_me = request.POST.get('remember_me')
        
        if not username or not password:
            messages.error(request, 'Please enter both username and password.')
            return render(request, 'auth/login.html')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                
                # Handle remember me functionality
                if not remember_me:
                    request.session.set_expiry(0)  # Session expires when browser closes
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect to next page or dashboard
                next_page = request.GET.get('next', 'dashboard')
                return redirect(next_page)
            else:
                messages.error(request, 'Your account has been deactivated. Please contact administrator.')
        else:
            messages.error(request, 'Invalid username or password. Please try again.')
    
    return render(request, 'auth/login.html')

@login_required
def logout_view(request):
    """
    Handle user logout
    """
    user_name = request.user.get_full_name() or request.user.username
    logout(request)
    messages.success(request, f'You have been logged out successfully. See you soon!')
    return redirect('login')



@login_required
def dashboard(request):
    """Main dashboard view"""
    # Get today's sales summary
    today = timezone.now().date()
    today_sales = Sale.objects.filter(created_at__date=today)
    
    # Calculate statistics
    total_sales_today = today_sales.aggregate(
        total=Sum('final_amount')
    )['total'] or 0
    
    total_transactions = today_sales.count()
    
    # Get recent sales
    recent_sales = Sale.objects.select_related('customer', 'cashier').order_by('-created_at')[:5]
    
    # Get low stock products
    low_stock_products = Product.objects.filter(
        stock_quantity__lte=models.F('min_stock_level'),
        is_active=True
    )[:5]
    
    # Get categories for sidebar
    categories = Category.objects.all()
    
    # Get featured/recent products
    products = Product.objects.filter(is_active=True).order_by('-created_at')[:8]
    
    context = {
        'total_sales_today': total_sales_today,
        'total_transactions': total_transactions,
        'recent_sales': recent_sales,
        'low_stock_products': low_stock_products,
        'categories': categories,
        'products': products,
        'current_date': timezone.now(),
    }
    
    return render(request, 'dashboard.html', context)



@login_required
def get_products_by_category(request, category_id):
    """AJAX view to get products by category"""
    products = Product.objects.filter(
        category_id=category_id, 
        is_active=True,
        stock_quantity__gt=0
    )
    
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'price': float(product.price),
            'stock': product.stock_quantity,
            'image': product.image.url if product.image else None,
        })
    
    return JsonResponse({'products': products_data})


from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from decimal import Decimal
import json
from datetime import datetime, timedelta
import uuid
import logging
import requests
import base64
from django.db import transaction
from django.conf import settings

from .models import (Product, Category, Sale, SaleItem, Customer, Discount, 
                    Inventory, Payment)  # Add Payment model

# Set up logging for M-Pesa debugging
logger = logging.getLogger(__name__)

# M-Pesa Service Class (copied from your ecommerce)
class MpesaService:
    def __init__(self):
        self.consumer_key = getattr(settings, 'MPESA_CONSUMER_KEY', '')
        self.consumer_secret = getattr(settings, 'MPESA_CONSUMER_SECRET', '')
        self.business_shortcode = getattr(settings, 'MPESA_BUSINESS_SHORTCODE', '')
        self.passkey = getattr(settings, 'MPESA_PASSKEY', '')
        self.environment = getattr(settings, 'MPESA_ENVIRONMENT', 'sandbox')
        
        logger.info(f"M-Pesa Service initialized - Environment: {self.environment}, Shortcode: {self.business_shortcode}")
        
        if not all([self.consumer_key, self.consumer_secret, self.business_shortcode, self.passkey]):
            logger.error("Missing M-Pesa configuration in settings")
            raise ValueError("M-Pesa configuration is incomplete. Check your settings.")
        
        if self.environment == 'sandbox':
            self.base_url = 'https://sandbox.safaricom.co.ke'
        else:
            self.base_url = 'https://api.safaricom.co.ke'
    
    def get_access_token(self):
        """Get M-Pesa access token with enhanced error handling"""
        logger.info("Requesting M-Pesa access token")
        
        try:
            url = f"{self.base_url}/oauth/v1/generate?grant_type=client_credentials"
            credentials = f"{self.consumer_key}:{self.consumer_secret}"
            encoded_credentials = base64.b64encode(credentials.encode()).decode()
            
            headers = {
                'Authorization': f'Basic {encoded_credentials}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(url, headers=headers, timeout=30)
            
            if response.status_code != 200:
                logger.error(f"Token request failed with status {response.status_code}")
                raise Exception(f"Token request failed: HTTP {response.status_code}")
            
            token_data = response.json()
            access_token = token_data.get('access_token')
            
            if not access_token:
                raise Exception("No access token received from API")
            
            logger.info("Access token obtained successfully")
            return access_token
            
        except Exception as e:
            logger.exception(f"Error getting access token: {str(e)}")
            raise Exception(f"Failed to get access token: {str(e)}")
    
    def generate_password(self):
        """Generate M-Pesa password"""
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        password_string = f"{self.business_shortcode}{self.passkey}{timestamp}"
        password = base64.b64encode(password_string.encode()).decode()
        return password, timestamp
    
    def stk_push(self, phone_number, amount, account_reference, transaction_desc):
        """Initiate STK Push with comprehensive error handling"""
        logger.info(f"Starting STK Push - Phone: {phone_number}, Amount: {amount}")
        
        try:
            access_token = self.get_access_token()
            password, timestamp = self.generate_password()
            
            url = f"{self.base_url}/mpesa/stkpush/v1/processrequest"
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            amount = int(float(amount))
            callback_url = getattr(settings, 'MPESA_CALLBACK_URL', '')
            
            payload = {
                'BusinessShortCode': self.business_shortcode,
                'Password': password,
                'Timestamp': timestamp,
                'TransactionType': 'CustomerPayBillOnline',
                'Amount': amount,
                'PartyA': phone_number,
                'PartyB': self.business_shortcode,
                'PhoneNumber': phone_number,
                'CallBackURL': callback_url,
                'AccountReference': account_reference,
                'TransactionDesc': transaction_desc
            }
            
            logger.info("Sending STK Push request...")
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            
            if response.status_code != 200:
                raise Exception(f"STK Push API error: HTTP {response.status_code}")
            
            response_data = response.json()
            response_code = response_data.get('ResponseCode')
            
            if response_code != '0':
                error_desc = response_data.get('ResponseDescription', 'Unknown error')
                raise Exception(f"STK Push failed: {error_desc}")
            
            logger.info("STK Push completed successfully")
            return response_data
            
        except Exception as e:
            logger.exception(f"STK Push failed: {str(e)}")
            raise Exception(f"STK Push failed: {str(e)}")


@login_required
def pos_terminal(request):
    """Enhanced POS Terminal with M-Pesa support"""
    categories = Category.objects.all()
    products = Product.objects.filter(is_active=True, stock_quantity__gt=0)
    customers = Customer.objects.all()
    discounts = Discount.objects.filter(is_active=True)
    
    context = {
        'categories': categories,
        'products': products,
        'customers': customers,
        'discounts': discounts,
    }
    
    return render(request, 'terminal.html', context)


@login_required
def process_sale(request):
    """Enhanced process_sale with M-Pesa integration"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Extract data
            cart_items = data.get('items', [])
            customer_id = data.get('customer_id')
            payment_method = data.get('payment_method', 'cash')
            amount_paid = Decimal(str(data.get('amount_paid', 0)))
            discount_id = data.get('discount_id')
            phone_number = data.get('phone_number', '')  # For M-Pesa
            
            if not cart_items:
                return JsonResponse({'success': False, 'error': 'No items in cart'})
            
            # Validate phone number for M-Pesa
            if payment_method == 'mpesa':
                cleaned_phone = clean_phone_number(phone_number)
                if not cleaned_phone:
                    return JsonResponse({
                        'success': False, 
                        'error': 'Valid phone number is required for M-Pesa payment'
                    })
                phone_number = cleaned_phone
            
            with transaction.atomic():
                # Calculate totals (same as before)
                total_amount = Decimal('0.00')
                sale_items = []
                
                for item in cart_items:
                    product = get_object_or_404(Product, id=item['product_id'])
                    quantity = int(item['quantity'])
                    unit_price = product.price
                    
                    # Check stock
                    if product.stock_quantity < quantity:
                        return JsonResponse({
                            'success': False, 
                            'error': f'Insufficient stock for {product.name}'
                        })
                    
                    item_total = unit_price * quantity
                    total_amount += item_total
                    
                    sale_items.append({
                        'product': product,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'total_price': item_total,
                    })
                
                # Apply discount
                discount_amount = Decimal('0.00')
                if discount_id:
                    discount = get_object_or_404(Discount, id=discount_id)
                    if discount.is_valid and total_amount >= discount.minimum_amount:
                        discount_amount = (total_amount * discount.percentage) / 100
                
                # Calculate final amount
                tax_amount = Decimal('0.00')
                final_amount = total_amount - discount_amount + tax_amount
                
                # Calculate change (only for cash)
                change_given = Decimal('0.00')
                if payment_method == 'cash':
                    change_given = amount_paid - final_amount if amount_paid >= final_amount else Decimal('0.00')
                else:
                    amount_paid = final_amount  # For non-cash payments
                
                # Create sale record
                sale = Sale.objects.create(
                    customer_id=customer_id if customer_id else None,
                    cashier=request.user,
                    total_amount=total_amount,
                    discount_amount=discount_amount,
                    tax_amount=tax_amount,
                    final_amount=final_amount,
                    payment_method=payment_method,
                    amount_paid=amount_paid,
                    change_given=change_given,
                    status='pending' if payment_method == 'mpesa' else 'completed'
                )
                
                logger.info(f"Sale created: {sale.sale_number}")
                
                # Create sale items and update inventory
                for item_data in sale_items:
                    SaleItem.objects.create(
                        sale=sale,
                        product=item_data['product'],
                        quantity=item_data['quantity'],
                        unit_price=item_data['unit_price'],
                        total_price=item_data['total_price'],
                    )
                    
                    # Update product stock (only if not M-Pesa or if M-Pesa payment succeeds)
                    if payment_method != 'mpesa':
                        product = item_data['product']
                        product.stock_quantity -= item_data['quantity']
                        product.save()
                        
                        # Create inventory transaction
                        Inventory.objects.create(
                            product=product,
                            transaction_type='sale',
                            quantity=-item_data['quantity'],
                            notes=f'Sale {sale.sale_number}',
                            user=request.user,
                        )
                
                # Update customer loyalty points (only if payment completes)
                if customer_id and payment_method != 'mpesa':
                    customer = Customer.objects.get(id=customer_id)
                    points_earned = int(final_amount / 10)
                    customer.loyalty_points += points_earned
                    customer.total_spent += final_amount
                    customer.save()
                
                # Handle M-Pesa payment
                if payment_method == 'mpesa':
                    logger.info(f"Initiating M-Pesa payment for sale {sale.sale_number}")
                    
                    try:
                        mpesa_service = MpesaService()
                        response = mpesa_service.stk_push(
                            phone_number=phone_number,
                            amount=int(final_amount),
                            account_reference=f"Sale-{sale.sale_number}",
                            transaction_desc=f"Payment for Sale {sale.sale_number}"
                        )
                        
                        if response.get('ResponseCode') == '0':
                            checkout_request_id = response.get('CheckoutRequestID')
                            
                            # Create Payment record
                            Payment.objects.create(
                                sale=sale,  # Link to sale instead of order
                                checkout_request_id=checkout_request_id,
                                status="PENDING",
                                raw_response=response,
                                phone_number=phone_number,
                                amount=final_amount
                            )
                            
                            logger.info(f"STK Push successful. CheckoutRequestID: {checkout_request_id}")
                            
                            return JsonResponse({
                                'success': True,
                                'payment_pending': True,
                                'message': 'STK Push sent to your phone. Please enter your M-Pesa PIN.',
                                'checkout_request_id': checkout_request_id,
                                'sale_id': sale.id,
                                'sale_number': sale.sale_number,
                                'total': float(final_amount)
                            })
                        else:
                            error_msg = response.get('ResponseDescription', 'Unknown error')
                            logger.error(f"STK Push failed: {error_msg}")
                            sale.delete()  # Clean up failed sale
                            
                            return JsonResponse({
                                'success': False,
                                'error': f"M-Pesa payment failed: {error_msg}"
                            })
                            
                    except Exception as e:
                        logger.exception(f"M-Pesa payment initialization failed: {str(e)}")
                        sale.delete()  # Clean up failed sale
                        
                        return JsonResponse({
                            'success': False,
                            'error': f"Payment initialization failed: {str(e)}"
                        })
                
                # For non-M-Pesa payments, return success immediately
                return JsonResponse({
                    'success': True,
                    'payment_pending': False,
                    'sale_id': sale.id,
                    'sale_number': sale.sale_number,
                    'total': float(final_amount),
                    'change': float(change_given),
                })
                
        except Exception as e:
            logger.exception(f"Error processing sale: {str(e)}")
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def clean_phone_number(phone):
    """Clean and format phone number for M-Pesa"""
    if not phone:
        return None
    
    # Remove all non-digit characters
    phone = ''.join(filter(str.isdigit, str(phone)))
    
    # Handle Kenyan phone numbers
    if phone.startswith('0'):
        phone = '254' + phone[1:]
    elif phone.startswith('+254'):
        phone = phone[1:]
    elif phone.startswith('254'):
        pass
    elif len(phone) == 9:
        phone = '254' + phone
    
    # Validate length
    if len(phone) == 12 and phone.startswith('254'):
        return phone
    
    return None


@csrf_exempt
@require_POST
def mpesa_callback(request):
    """Handle M-Pesa callback for POS transactions"""
    logger.info("M-Pesa callback received for POS")
    logger.debug(f"Callback request body: {request.body}")
    
    try:
        callback_data = json.loads(request.body)
        stk_callback = callback_data.get('Body', {}).get('stkCallback', {})
        result_code = stk_callback.get('ResultCode')
        checkout_request_id = stk_callback.get('CheckoutRequestID')
        result_desc = stk_callback.get('ResultDesc')
        
        logger.info(f"Callback details - ResultCode: {result_code}, CheckoutRequestID: {checkout_request_id}")
        
        # Find Payment by checkout_request_id
        payment = Payment.objects.filter(checkout_request_id=checkout_request_id).first()
        if not payment:
            logger.error(f"No Payment found for CheckoutRequestID: {checkout_request_id}")
            return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})
        
        sale = payment.sale
        
        if result_code == 0:  # Success
            callback_metadata = stk_callback.get('CallbackMetadata', {}).get('Item', [])
            transaction_data = {item['Name']: item.get('Value') for item in callback_metadata}
            
            logger.info(f"Transaction data: {transaction_data}")
            
            # Update payment record
            payment.status = "SUCCESS"
            payment.mpesa_receipt = transaction_data.get('MpesaReceiptNumber')
            payment.phone_number = transaction_data.get('PhoneNumber')
            payment.transaction_date = str(transaction_data.get('TransactionDate'))
            payment.amount = transaction_data.get('Amount')
            payment.raw_response = callback_data
            payment.save()
            
            # Update sale status and complete inventory updates
            sale.status = "completed"
            sale.save()
            
            # Now update inventory (was deferred for M-Pesa)
            for sale_item in sale.items.all():
                product = sale_item.product
                product.stock_quantity -= sale_item.quantity
                product.save()
                
                # Create inventory transaction
                Inventory.objects.create(
                    product=product,
                    transaction_type='sale',
                    quantity=-sale_item.quantity,
                    notes=f'Sale {sale.sale_number} - M-Pesa Payment Confirmed',
                    user=sale.cashier,
                )
            
            # Update customer loyalty points if applicable
            if sale.customer:
                customer = sale.customer
                points_earned = int(sale.final_amount / 10)
                customer.loyalty_points += points_earned
                customer.total_spent += sale.final_amount
                customer.save()
            
            logger.info(f"Payment completed - Receipt: {payment.mpesa_receipt}")
        
        else:  # Failed
            payment.status = "FAILED"
            payment.raw_response = callback_data
            payment.save()
            
            sale.status = "cancelled"
            sale.save()
            
            logger.error(f"Payment failed - ResultCode: {result_code}, ResultDesc: {result_desc}")
        
        return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})
    
    except Exception as e:
        logger.exception(f"Error processing callback: {str(e)}")
        return JsonResponse({'ResultCode': 1, 'ResultDesc': f'Error processing callback: {str(e)}'})


@login_required
def check_payment_status(request):
    """Check M-Pesa payment status via AJAX polling"""
    checkout_request_id = request.GET.get('checkout_request_id')
    logger.info(f"Checking payment status for CheckoutRequestID: {checkout_request_id}")
    
    if not checkout_request_id:
        return JsonResponse({'error': 'Checkout request ID required'}, status=400)
    
    try:
        payment = Payment.objects.filter(checkout_request_id=checkout_request_id).first()
        
        if payment:
            if payment.status == "SUCCESS":
                return JsonResponse({
                    'status': 'SUCCESS',
                    'message': 'Payment completed successfully!',
                    'sale_number': payment.sale.sale_number,
                    'mpesa_receipt': payment.mpesa_receipt
                })
            elif payment.status == "FAILED":
                return JsonResponse({
                    'status': 'FAILED',
                    'message': 'Payment failed. Please try again.'
                })
            else:
                return JsonResponse({
                    'status': 'PENDING',
                    'message': 'Please complete the payment on your phone...'
                })
        
        # Default pending response
        return JsonResponse({
            'status': 'PENDING',
            'message': 'Payment is being processed...'
        })
        
    except Exception as e:
        logger.exception(f"Error checking payment status: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def sale_detail(request, sale_id):
    """Enhanced sale detail view with payment info"""
    sale = get_object_or_404(Sale, id=sale_id)
    sale_items = sale.items.select_related('product')
    
    # Get payment info if M-Pesa
    payment = None
    if sale.payment_method == 'mpesa':
        payment = Payment.objects.filter(sale=sale).first()
    
    context = {
        'sale': sale,
        'sale_items': sale_items,
        'payment': payment,
    }
    
    return render(request, 'sale_detail.html', context)


# Additional utility functions
@login_required
def get_pending_mpesa_sales(request):
    """Get sales with pending M-Pesa payments"""
    pending_sales = Sale.objects.filter(
        payment_method='mpesa',
        status='pending'
    ).select_related('customer', 'cashier')
    
    sales_data = []
    for sale in pending_sales:
        payment = Payment.objects.filter(sale=sale).first()
        sales_data.append({
            'id': sale.id,
            'sale_number': sale.sale_number,
            'total': float(sale.final_amount),
            'customer': sale.customer.name if sale.customer else 'Walk-in',
            'created_at': sale.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'checkout_request_id': payment.checkout_request_id if payment else None,
            'phone_number': payment.phone_number if payment else None
        })
    
    return JsonResponse({'pending_sales': sales_data})

@login_required
def search_products(request):
    """AJAX view to search products"""
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'products': []})
    
    products = Product.objects.filter(
        Q(name__icontains=query) | Q(sku__icontains=query),
        is_active=True,
        stock_quantity__gt=0
    )[:10]
    
    products_data = []
    for product in products:
        products_data.append({
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'price': float(product.price),
            'stock': product.stock_quantity,
            'category': product.category.name,
        })
    
    return JsonResponse({'products': products_data})


@login_required
def sales_history(request):
    """View sales history"""
    sales = Sale.objects.select_related('customer', 'cashier').order_by('-created_at')
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        sales = sales.filter(created_at__date__gte=date_from)
    if date_to:
        sales = sales.filter(created_at__date__lte=date_to)
    
    context = {
        'sales': sales,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'sales_history.html', context)



@login_required
def inventory_management(request):
    """Inventory management view"""
    products = Product.objects.select_related('category').all()
    categories = Category.objects.all()
    
    # Filter by category if provided
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) | 
            Q(sku__icontains=search_query)
        )
    
    context = {
        'products': products,
        'categories': categories,
        'selected_category': category_id,
        'search_query': search_query,
    }
    
    return render(request, 'inventory.html', context)

@login_required
@require_http_methods(["POST"])
def update_inventory(request):
    """AJAX view to update inventory stock"""
    try:
        data = json.loads(request.body)
        product_id = data.get('product_id')
        transaction_type = data.get('transaction_type')
        quantity = int(data.get('quantity', 0))
        notes = data.get('notes', '')
        
        product = get_object_or_404(Product, id=product_id, is_active=True)
        
        # Validate stock out doesn't exceed available stock
        if transaction_type == 'out' and abs(quantity) > product.stock_quantity:
            return JsonResponse({
                'success': False,
                'error': f'Insufficient stock. Available: {product.stock_quantity}'
            })
        
        # Update product stock
        if transaction_type == 'in':
            product.stock_quantity += abs(quantity)
            inventory_quantity = abs(quantity)
        else:  # out
            product.stock_quantity -= abs(quantity)
            inventory_quantity = -abs(quantity)
        
        product.save()
        
        # Create inventory transaction record
        Inventory.objects.create(
            product=product,
            transaction_type=transaction_type,
            quantity=inventory_quantity,
            notes=notes,
            user=request.user
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Inventory updated successfully',
            'new_stock': product.stock_quantity,
            'is_low_stock': product.is_low_stock
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def export_inventory(request):
    """Export inventory to Excel"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Define headers
        headers = [
            'SKU', 'Product Name', 'Category', 'Current Stock', 
            'Min Level', 'Unit Price', 'Cost Price', 'Stock Value', 'Status'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get products data
        products = Product.objects.select_related('category').filter(is_active=True).order_by('name')
        
        # Add data rows
        for row, product in enumerate(products, 2):
            stock_value = product.stock_quantity * product.cost_price
            
            # Determine status
            if product.stock_quantity == 0:
                status = "Out of Stock"
            elif product.is_low_stock:
                status = "Low Stock"
            else:
                status = "In Stock"
            
            row_data = [
                product.sku,
                product.name,
                product.category.name,
                product.stock_quantity,
                product.min_stock_level,
                float(product.price),
                float(product.cost_price),
                float(stock_value),
                status
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Export failed: {str(e)}'
        })

@login_required
@require_http_methods(["POST"])
def add_product(request):
    """AJAX view to add new product"""
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['name', 'category_id', 'sku', 'price']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'error': f'{field.replace("_", " ").title()} is required'
                })
        
        # Check if SKU already exists
        if Product.objects.filter(sku=data['sku']).exists():
            return JsonResponse({
                'success': False,
                'error': 'SKU already exists'
            })
        
        # Get category
        try:
            category = Category.objects.get(id=data['category_id'])
        except Category.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Invalid category'
            })
        
        # Create product
        product = Product.objects.create(
            name=data['name'],
            category=category,
            sku=data['sku'],
            description=data.get('description', ''),
            price=data['price'],
            cost_price=data.get('cost_price', 0),
            stock_quantity=data.get('stock_quantity', 0),
            min_stock_level=data.get('min_stock_level', 5),
            is_active=True
        )
        
        # If initial stock is added, create inventory record
        if product.stock_quantity > 0:
            Inventory.objects.create(
                product=product,
                transaction_type='in',
                quantity=product.stock_quantity,
                notes='Initial stock',
                user=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Product added successfully',
            'product_id': product.id,
            'product_name': product.name
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def get_product_history(request, product_id):
    """Get inventory history for a specific product"""
    try:
        product = get_object_or_404(Product, id=product_id)
        
        history = Inventory.objects.filter(product=product).select_related('user').order_by('-created_at')[:50]
        
        history_data = []
        for record in history:
            history_data.append({
                'date': record.created_at.strftime('%Y-%m-%d %H:%M'),
                'type': record.get_transaction_type_display(),
                'quantity': record.quantity,
                'notes': record.notes,
                'user': record.user.username if record.user else 'System'
            })
        
        return JsonResponse({
            'success': True,
            'product_name': product.name,
            'current_stock': product.stock_quantity,
            'history': history_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q
from django.db.models.functions import Extract
from django.utils import timezone
from datetime import datetime, timedelta
import json
from decimal import Decimal
from .models import Customer, Sale, SaleItem, Product, Category

# Customer Views
@login_required
def customers_list(request):
    """Display customers list page"""
    return render(request, 'customers.html')

@login_required
def customers_ajax(request):
    """AJAX endpoint for customer data"""
    if request.method == 'GET':
        customers = Customer.objects.all().order_by('-created_at')
        data = []
        for customer in customers:
            data.append({
                'id': customer.id,
                'name': customer.name,
                'email': customer.email,
                'phone': customer.phone,
                'address': customer.address,
                'loyalty_tier': customer.get_loyalty_tier_display(),
                'loyalty_points': customer.loyalty_points,
                'total_spent': str(customer.total_spent),
                'discount_percentage': customer.discount_percentage,
                'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M')
            })
        return JsonResponse({'data': data})

@login_required
def customer_detail_ajax(request, customer_id):
    """Get single customer details"""
    try:
        customer = Customer.objects.get(id=customer_id)
        data = {
            'id': customer.id,
            'name': customer.name,
            'email': customer.email,
            'phone': customer.phone,
            'address': customer.address,
            'loyalty_tier': customer.loyalty_tier,
            'loyalty_points': customer.loyalty_points,
            'total_spent': str(customer.total_spent),
            'created_at': customer.created_at.strftime('%Y-%m-%d %H:%M')
        }
        return JsonResponse({'success': True, 'data': data})
    except Customer.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Customer not found'})

@csrf_exempt
@login_required
def customer_create_ajax(request):
    """Create new customer via AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            customer = Customer.objects.create(
                name=data.get('name'),
                email=data.get('email', ''),
                phone=data.get('phone', ''),
                address=data.get('address', ''),
                loyalty_tier=data.get('loyalty_tier', 'bronze'),
                loyalty_points=int(data.get('loyalty_points', 0))
            )
            return JsonResponse({
                'success': True, 
                'message': 'Customer created successfully',
                'data': {
                    'id': customer.id,
                    'name': customer.name,
                    'email': customer.email,
                    'phone': customer.phone
                }
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@login_required
def customer_update_ajax(request, customer_id):
    """Update customer via AJAX"""
    if request.method == 'POST':
        try:
            customer = Customer.objects.get(id=customer_id)
            data = json.loads(request.body)
            
            customer.name = data.get('name', customer.name)
            customer.email = data.get('email', customer.email)
            customer.phone = data.get('phone', customer.phone)
            customer.address = data.get('address', customer.address)
            customer.loyalty_tier = data.get('loyalty_tier', customer.loyalty_tier)
            customer.loyalty_points = int(data.get('loyalty_points', customer.loyalty_points))
            customer.save()
            
            return JsonResponse({'success': True, 'message': 'Customer updated successfully'})
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Customer not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@login_required
def customer_delete_ajax(request, customer_id):
    """Delete customer via AJAX"""
    if request.method == 'POST':
        try:
            customer = Customer.objects.get(id=customer_id)
            customer.delete()
            return JsonResponse({'success': True, 'message': 'Customer deleted successfully'})
        except Customer.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Customer not found'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

# Reports Views
@login_required
def reports_view(request):
    """Display reports dashboard"""
    return render(request, 'reports.html')

@login_required
def sales_data_ajax(request):
    """Get sales data for the past year"""
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)
    
    # Monthly sales data
    monthly_sales = Sale.objects.filter(
        created_at__range=[start_date, end_date]
    ).extra(
        select={'month': "DATE_TRUNC('month', created_at)"}
    ).values('month').annotate(
        total_sales=Sum('final_amount'),
        total_orders=Count('id')
    ).order_by('month')
    
    # Format data for Chart.js
    months = []
    sales = []
    orders = []
    
    for item in monthly_sales:
        months.append(item['month'].strftime('%Y-%m'))
        sales.append(float(item['total_sales'] or 0))
        orders.append(item['total_orders'])
    
    return JsonResponse({
        'months': months,
        'sales': sales,
        'orders': orders
    })

@login_required
def top_products_ajax(request):
    """Get most sold products"""
    top_products = SaleItem.objects.values(
        'product__name'
    ).annotate(
        total_quantity=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_quantity')[:10]
    
    products = []
    quantities = []
    revenues = []
    
    for item in top_products:
        products.append(item['product__name'])
        quantities.append(item['total_quantity'])
        revenues.append(float(item['total_revenue']))
    
    return JsonResponse({
        'products': products,
        'quantities': quantities,
        'revenues': revenues
    })

@login_required
def daily_sales_analysis_ajax(request):
    """Analyze sales by day of week"""
    sales_by_day = Sale.objects.annotate(
        day_of_week=Extract('created_at', 'dow')
    ).values('day_of_week').annotate(
        total_sales=Sum('final_amount'),
        total_orders=Count('id'),
        avg_order_value=Sum('final_amount') / Count('id')
    ).order_by('day_of_week')
    
    days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
    day_sales = [0] * 7
    day_orders = [0] * 7
    day_avg = [0] * 7
    
    for item in sales_by_day:
        day_index = item['day_of_week']
        day_sales[day_index] = float(item['total_sales'] or 0)
        day_orders[day_index] = item['total_orders']
        day_avg[day_index] = float(item['avg_order_value'] or 0)
    
    return JsonResponse({
        'days': days,
        'sales': day_sales,
        'orders': day_orders,
        'avg_values': day_avg
    })

@login_required
def sales_summary_ajax(request):
    """Get sales summary statistics"""
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    last_month_start = (this_month_start - timedelta(days=1)).replace(day=1)
    
    # Today's sales
    today_sales = Sale.objects.filter(
        created_at__date=today
    ).aggregate(
        total_sales=Sum('final_amount'),
        total_orders=Count('id')
    )
    
    # This month's sales
    this_month_sales = Sale.objects.filter(
        created_at__date__gte=this_month_start
    ).aggregate(
        total_sales=Sum('final_amount'),
        total_orders=Count('id')
    )
    
    # Last month's sales
    last_month_sales = Sale.objects.filter(
        created_at__date__gte=last_month_start,
        created_at__date__lt=this_month_start
    ).aggregate(
        total_sales=Sum('final_amount'),
        total_orders=Count('id')
    )
    
    # Top customer
    top_customer = Customer.objects.order_by('-total_spent').first()
    
    return JsonResponse({
        'today_sales': float(today_sales['total_sales'] or 0),
        'today_orders': today_sales['total_orders'] or 0,
        'month_sales': float(this_month_sales['total_sales'] or 0),
        'month_orders': this_month_sales['total_orders'] or 0,
        'last_month_sales': float(last_month_sales['total_sales'] or 0),
        'top_customer': top_customer.name if top_customer else 'No customers',
        'total_customers': Customer.objects.count(),
        'total_products': Product.objects.count(),
        'low_stock_products': Product.objects.filter(
            stock_quantity__lte=F('min_stock_level')
        ).count()
    })

# Settings Views
@login_required
def settings_view(request):
    """Display settings page"""
    return render(request, 'settings.html')